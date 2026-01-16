
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO
from typing import List, Dict

class OKRSheetGenerator:
    """Generates OKR Evaluation Excel Sheet"""
    
    def __init__(self):
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid") # Blue
        self.section_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid") # Yellow
        self.header_font = Font(name='Times New Roman', size=11, bold=True, color="FFFFFF")
        self.section_font = Font(name='Times New Roman', size=11, bold=True, italic=True)
        self.item_font_bold_italic = Font(name='Times New Roman', size=11, bold=True, italic=True)
        self.normal_font = Font(name='Times New Roman', size=11)
        self.thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def generate_excel(self, users_data: List[Dict], cycle_name: str) -> BytesIO:
        """
        Generate Excel file with user data.
        users_data: List of dicts, each containing 'name', 'stats' (derived from get_user_excel_data)
        """
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Đánh giá OKRs"

        # 1. Setup Columns
        ws.column_dimensions['A'].width = 5   # TT
        ws.column_dimensions['B'].width = 60  # Noi dung
        ws.column_dimensions['C'].width = 10  # Tu cham diem
        
        # User columns
        user_names = [u['name'] for u in users_data]
        if not user_names:
             user_names = ["Demo User"] # Fallback if empty to avoid index error in headers

        for i, user in enumerate(user_names):
            col_letter = get_column_letter(4 + i)
            ws.column_dimensions[col_letter].width = 15
            ws.cell(row=2, column=4 + i).value = user

        # 2. Header
        # Avoid error if no users
        end_col = 3 + len(user_names) if user_names else 4
        ws.merge_cells(f'D1:{get_column_letter(end_col)}1')
        title_cell = ws['D1']
        title_cell.value = f"ĐÁNH GIÁ OKRs - {cycle_name}".upper()
        title_cell.font = Font(name='Times New Roman', size=14, bold=True)
        title_cell.alignment = self.center_align

        # Row 2 Headers
        headers = ["TT", "Nội dung", "Tự chấm điểm"]
        for col, text in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col)
            cell.value = text
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # Apply header style to user columns
        for i in range(len(user_names)):
            cell = ws.cell(row=2, column=4 + i)
            cell.fill = self.header_fill
            cell.font = self.header_font
            cell.alignment = self.center_align
            cell.border = self.thin_border

        # 3. Content Template & Data Filling
        # Structure: (TT, Content, Score/Ref, Type, Key_for_filling)
        # Type: section, item, subitem, score_row
        # Key_for_filling: Identifying which stat maps to this row
        
        data_structure = [
            ("I", "CHẤT LƯỢNG THỰC THI OKR", "", "section", None),
            ("1", "Kết quả thực tế so với mục tiêu: (Dịch chuyển so với tháng trước/33,3%)", "", "item", "okr_shift_display"),
            ("-", "Nhỏ hơn 25%", 5, "score_row", "shift_lt_25"),
            ("-", "Từ 25 - 50%", 10, "score_row", "shift_25_50"),
            ("-", "Từ 50 - 75%", 15, "score_row", "shift_50_75"),
            ("-", "Từ 75 - 100%", 18, "score_row", "shift_75_100"),
            ("-", "Trên 100% hoặc có đột phá lớn (Cấp trên ghi nhận)", 20, "score_row", "shift_gt_100"),
            
            ("2", "Tiến độ và tính kỷ luật", "", "item", None),
            ("2.1", "Đầy đủ OKRs cá nhân được cập nhật trên Base Goal", "Yes/No", "subitem", "has_okrs"),
            ("-", "Có Check-in trên base hàng tuần (2 điểm/lần check-in)", 8, "score_row", "checkin_score"),
            ("-", "Có check-in với người khác, cấp quản lý, làm việc chung OKRs trong bộ phận", 2, "score_row", "collab_score"), 
            
            ("2.2", "Chất lượng check - in và hành động trong KR", "", "subitem", None),
            ("-", "Không có hành động rõ ràng", 1, "score_row", "quality_low"),
            ("-", "Chỉ báo cáo trạng thái (đang làm, đang cố, ...)", 3, "score_row", "quality_med"),
            ("-", "Có mô tả hành động rõ ràng cụ thể + hướng giải quyết", 5, "score_row", "quality_high"),
            
            ("II", "TẦM QUAN TRỌNG CỦA OKR", "", "section", None),
            ("1", "Mức độ đóng góp vào mục tiêu công ty", "", "item", None),
            ("-", "Mục tiêu mang tính nội bộ/cá nhân", 1, "score_row", "align_personal"),
            ("-", "Mục tiêu hỗ trợ gián tiếp Doanh thu/Khách hàng/Chất lượng (1)", 2, "score_row", "align_indirect_1"),
            ("-", "Mục tiêu hỗ trợ gián tiếp Doanh thu/Khách hàng/Chất lượng (2)", 3, "score_row", "align_indirect_2"),
            ("-", "Mục tiêu liên quan trực tiếp Doanh thu/Khách hàng/Chất lượng", 4, "score_row", "align_direct_1"),
            ("-", "Mục tiêu liên quan trực tiếp Doanh thu/Khách hàng/Chất lượng (High)", 5, "score_row", "align_direct_2"),
            
            ("2", "Mức độ ưu tiên mục tiêu của Quý", "", "item", None),
            ("-", "Bình thường", 1, "score_row", "prio_normal"),
            ("-", "Quan trọng", 2, "score_row", "prio_important_1"),
            ("-", "Quan trọng (High)", 3, "score_row", "prio_important_2"),
            ("-", "Rất quan trọng", 4, "score_row", "prio_very_important_1"),
            ("-", "Rất quan trọng (High)", 5, "score_row", "prio_very_important_2"),
            
            ("3", "Tính khó/tầm ảnh hưởng đến hệ thống", "", "item", None),
            ("-", "Tác động với cá nhân", 1, "score_row", "impact_personal"),
            ("-", "Tác động nội bộ phòng ban/đội nhóm", 2, "score_row", "impact_team_1"),
            ("-", "Tác động nội bộ phòng ban/đội nhóm (High)", 3, "score_row", "impact_team_2"),
            ("-", "Tác động nhiều phòng ban/cả công ty", 4, "score_row", "impact_company_1"),
            ("-", "Tác động nhiều phòng ban/cả công ty (High)", 5, "score_row", "impact_company_2"),
        ]

        current_row = 3
        
        for tt, content, score_ref, style_type, key in data_structure:
            # A: TT
            cell_tt = ws.cell(row=current_row, column=1, value=tt)
            cell_tt.alignment = self.center_align
            cell_tt.border = self.thin_border
            
            # B: Content
            cell_content = ws.cell(row=current_row, column=2, value=content)
            cell_content.border = self.thin_border
            cell_content.alignment = self.left_align
            
            # C: Score
            cell_score = ws.cell(row=current_row, column=3, value=score_ref)
            cell_score.alignment = self.center_align
            cell_score.border = self.thin_border

            # User columns
            for i, user_data in enumerate(users_data):
                col_idx = 4 + i
                cell = ws.cell(row=current_row, column=col_idx)
                cell.border = self.thin_border
                cell.alignment = self.center_align
                
                stats = user_data.get('stats', {})
                
                if key: # Modified to check key for all types
                     if style_type == "item" and key == "okr_shift_display":
                         val = stats.get(key, '')
                         if val: 
                            # Display format: e.g. "35%"
                            cell.value = val
                     elif style_type == "score_row" or style_type == "subitem":
                         if key == "checkin_score":
                             val = stats.get('checkin_score_val', '')
                             if val: cell.value = val
                         elif key == "has_okrs":
                             val = stats.get('has_okrs', '')
                             cell.value = val 
                             if val == 'No': cell.font = Font(color="FF0000")
                         elif key in stats and stats[key]:
                             if isinstance(score_ref, (int, float)):
                                cell.value = score_ref 
                             else:
                                cell.value = "x"

            # Styling
            if style_type == "section":
                for col in range(1, 4 + len(user_names)):
                    ws.cell(row=current_row, column=col).fill = self.section_fill
                cell_tt.font = self.section_font
                cell_content.font = self.section_font
                
            elif style_type == "item":
                cell_tt.font = self.item_font_bold_italic
                cell_content.font = self.item_font_bold_italic
                
            elif style_type == "subitem":
                 cell_content.font = Font(name='Times New Roman', size=11, bold=True)
                 if score_ref == "Yes/No":
                     cell_score.font = Font(name='Times New Roman', size=11, color="FF0000")
                     
            else: # score_row
                cell_content.font = self.normal_font

            current_row += 1

        # Total Row
        total_row_idx = current_row
        ws.cell(row=total_row_idx, column=2, value="Tổng cộng OKR").font = Font(name='Times New Roman', size=11, bold=True)
        ws.cell(row=total_row_idx, column=1).border = self.thin_border
        ws.cell(row=total_row_idx, column=2).border = self.thin_border
        ws.cell(row=total_row_idx, column=3).border = self.thin_border

        for i, _ in enumerate(user_names):
            col_idx = 4 + i
            cell = ws.cell(row=total_row_idx, column=col_idx)
            cell.fill = self.section_fill
            cell.font = Font(name='Times New Roman', size=11, bold=True)
            cell.border = self.thin_border
            # SUM formula
            col_letter = get_column_letter(col_idx)
            cell.value = f"=SUM({col_letter}3:{col_letter}{total_row_idx-1})"

        # 4. Legend (Fixed at B40-B44 as requested)
        legend_data = [
            (40, "Mức 1 ≤ 15", "FF0000"),       # Red
            (41, "15 < Mức 2 ≤ 25", "00B0F0"),  # Cyan/Blue
            (42, "25 < Mức 3 ≤ 35", "FFFF00"),  # Yellow
            (43, "35 < Mức 4 ≤ 45", "00B050"),  # Green
            (44, "Mức 5 > 45", "7030A0")        # Purple
        ]

        for r, text, color in legend_data:
            cell = ws.cell(row=r, column=2)
            cell.value = text
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
            cell.font = Font(name='Times New Roman', size=11, bold=True)
            cell.alignment = self.center_align
            # Merge cell B with C for better look? User said B40-44. Let's stick to B.
            # But usually legends might span. The image looks like it spans B to... well image is cut off.
            # But let's assume just B. Note column B width is 60 so it fits.
            cell.border = self.thin_border
            
            # Add border to all columns in this row for B to end_col? Use image as guide.
            # Image shows the colored bar spans the width of column B ("Nội dung").
            # It doesn't look like it spans "Tu cham diem" (C).
            # So applying to Column B only is correct.

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer
